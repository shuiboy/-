import re
import requests
# import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import *

lw = load_workbook('网易新闻.xlsx')
tb = lw['表1']
# tb.column
print(tb.max_column, tb.max_row)
tb.insert_rows(1)
a = ['标题', '栏目', '详情', '图片', '来源', '详情页', '其他1', '其他2', '其他3']
for w in range(1,10):
    tb.cell(1,w,a[w-1])
# tb.append(['标题', '栏目', '详情', '图片', '来源', '详情页', '其他1', '其他2', '其他3'])

# for i in range(1, tb.max_row):
#     for j in range(1, tb.max_column):
#         print(tb.cell(i, j).value, end=' ')
#     print()
#
tb_all = tb.values
for i in tb_all:
    print(i)
lw.save('网易新闻.xlsx')
lw.close()
